aimport tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import queue
import sys
import os
import requests
import re
import openpyxl
from bs4 import BeautifulSoup
from typing import Dict, List, Optional, Union, Set, Tuple
from decimal import Decimal
from datetime import datetime, timedelta
import warnings
warnings.filterwarnings('ignore')

# 日志级别常量
LOG_LEVEL = {
    'LOG': 1,    # 所有信息
    'OPER': 2,   # 操作信息和错误信息
    'ERR': 3     # 只输出错误信息
}

class FundDataFetcher:
    def __init__(self):
        self.request_headers: Dict[str, str] = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
    
    def fetch_fund_history_data(self, fund_code: str, target_date: str = None) -> Tuple[bool, Optional[Dict[str, Union[str, float]]]]:
        """获取基金指定日期的历史净值数据"""
        api_url = f"http://api.fund.eastmoney.com/f10/lsjz?fundCode={fund_code}&pageIndex=1&pageSize=30"
        
        # 添加必要的请求头
        headers = self.request_headers.copy()
        headers['Referer'] = 'http://fundf10.eastmoney.com/'
        
        try:
            response = requests.get(api_url, headers=headers, timeout=10)
            
            if response.status_code == 200:
                data = response.json()
                
                if data.get('Data') and data['Data'].get('LSJZList'):
                    fund_list = data['Data']['LSJZList']
                    
                    if target_date:
                        # 查找指定日期的数据
                        for fund_data in fund_list:
                            if fund_data.get('FSRQ') == target_date:
                                net_value = float(fund_data.get('DWJZ', 0))
                                change_percent = float(fund_data.get('JZZZL', 0))
                                return True, {
                                    'date': target_date,
                                    'net_value': net_value,
                                    'change_percent': change_percent
                                }
                        return True, None  # 指定日期无数据
                    else:
                        # 获取最新数据
                        latest_data = fund_list[0]
                        net_value = float(latest_data.get('DWJZ', 0))
                        change_percent = float(latest_data.get('JZZZL', 0))
                        date_str = latest_data.get('FSRQ', '')
                        
                        return True, {
                            'date': date_str,
                            'net_value': net_value,
                            'change_percent': change_percent
                        }
                else:
                    return False, None
            else:
                return False, None
                
        except Exception as e:
            return False, None

class FundExcelUpdater:
    def __init__(self, excel_file: str = "fund.xlsx"):
        self.excel_file_path: str = excel_file
        self.sheet_names: List[str] = ["美指净值", "债基净值", "A股净值"]
        self.excel_fund_values: Dict[str, Tuple[float, str]] = {}
    
    def read_excel_fund_values(self) -> None:
        """从Excel文件读取基金代码和净值，存储到字典中"""
        self.excel_fund_values.clear()
        
        try:
            workbook = openpyxl.load_workbook(self.excel_file_path)
            
            for sheet_name in self.sheet_names:
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    
                    for row in range(2, sheet.max_row + 1):
                        # 获取第二列的基金代码（与fund.py保持一致）
                        fund_code = sheet.cell(row=row, column=2).value
                        
                        if fund_code:
                            # 确保基金代码是字符串格式
                            fund_code_str = str(fund_code).strip()
                            
                            # 如果基金代码以'A'开头，去掉'A'前缀（与fund.py保持一致）
                            if fund_code_str.startswith('A'):
                                fund_code_str = fund_code_str[1:]
                            
                            # 验证基金代码格式，允许字母和数字
                            if not re.match(r'^[A-Za-z0-9]+$', fund_code_str):
                                continue
                            
                            # 获取第四列的净值和第五列的涨跌幅（与fund.py保持一致）
                            net_value = sheet.cell(row=row, column=4).value
                            percentage = sheet.cell(row=row, column=5).value
                            
                            if net_value is not None:
                                # 处理净值
                                converted_value = 0.0
                                try:
                                    if isinstance(net_value, (int, float, Decimal)):
                                        converted_value = float(net_value)
                                    elif isinstance(net_value, str):
                                        converted_value = float(net_value.strip())
                                except (ValueError, TypeError):
                                    converted_value = 0.0
                                
                                # 处理涨跌幅
                                percentage_str = "N/A"
                                if percentage is not None:
                                    if isinstance(percentage, str):
                                        percentage_str = percentage.strip()
                                    elif isinstance(percentage, (int, float, Decimal)):
                                        percentage_str = f"{float(percentage) * 100:.2f}%"
                                
                                self.excel_fund_values[fund_code_str] = (converted_value, percentage_str)
            
            workbook.close()
            
        except Exception as e:
            raise Exception(f"读取Excel文件失败: {str(e)}")
    
    def update_excel_values(self) -> None:
        """将更新后的基金数据写回Excel文件"""
        try:
            workbook = openpyxl.load_workbook(self.excel_file_path)
            
            for sheet_name in self.sheet_names:
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    
                    for row in range(2, sheet.max_row + 1):
                        # 获取第二列的基金代码（与fund.py保持一致）
                        fund_code = sheet.cell(row=row, column=2).value
                        
                        if fund_code:
                            fund_code_str = str(fund_code).strip()
                            
                            if fund_code_str in self.excel_fund_values:
                                net_value, percentage = self.excel_fund_values[fund_code_str]
                                
                                # 更新净值（第四列）
                                sheet.cell(row=row, column=4).value = net_value
                                
                                # 更新百分比（第五列）
                                if percentage and isinstance(percentage, str) and percentage.endswith('%'):
                                    sheet.cell(row=row, column=5).value = percentage
                                else:
                                    sheet.cell(row=row, column=5).value = f"{float(percentage):.2f}%"
            
            workbook.save(self.excel_file_path)
            workbook.close()
            
        except Exception as e:
            raise Exception(f"更新Excel文件失败: {str(e)}")

class FundUpdaterGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("基金数据更新工具")
        self.root.geometry("800x600")
        
        # 创建队列用于线程间通信
        self.log_queue = queue.Queue()
        
        # 默认日志级别
        self.log_level = LOG_LEVEL['LOG']  # 默认显示所有信息
        
        # 初始化组件
        self.setup_ui()
        
        # 启动日志处理循环
        self.process_log_queue()
        
        # 初始化基金数据处理器
        self.fund_data_fetcher = FundDataFetcher()
        self.excel_updater = None
        
        # 线程控制
        self.is_running = False
        self.update_thread = None
    
    def setup_ui(self):
        """设置用户界面"""
        # 主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 配置网格权重
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(3, weight=1)
        
        # 文件选择区域
        file_frame = ttk.LabelFrame(main_frame, text="文件设置", padding="5")
        file_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        file_frame.columnconfigure(1, weight=1)
        
        ttk.Label(file_frame, text="Excel文件:").grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        self.file_path_var = tk.StringVar(value="fund.xlsx")
        self.file_entry = ttk.Entry(file_frame, textvariable=self.file_path_var, state='readonly')
        self.file_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 5))
        
        self.browse_btn = ttk.Button(file_frame, text="浏览", command=self.browse_file)
        self.browse_btn.grid(row=0, column=2)
        
        # 控制按钮区域
        control_frame = ttk.Frame(main_frame)
        control_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.start_btn = ttk.Button(control_frame, text="开始更新", command=self.start_update)
        self.start_btn.grid(row=0, column=0, padx=(0, 10))
        
        self.stop_btn = ttk.Button(control_frame, text="停止", command=self.stop_update, state='disabled')
        self.stop_btn.grid(row=0, column=1, padx=(0, 10))
        
        self.reset_btn = ttk.Button(control_frame, text="重置", command=self.reset_ui)
        self.reset_btn.grid(row=0, column=2)
        
        # 日志级别选择
        ttk.Label(control_frame, text="日志级别:").grid(row=0, column=3, padx=(20, 5))
        self.log_level_var = tk.StringVar(value="所有信息")
        self.log_level_combo = ttk.Combobox(control_frame, textvariable=self.log_level_var, 
                                           values=["所有信息", "操作信息", "仅错误"], 
                                           state="readonly", width=10)
        self.log_level_combo.grid(row=0, column=4, padx=(0, 10))
        self.log_level_combo.bind("<<ComboboxSelected>>", self.on_log_level_changed)
        
        # 进度显示区域
        progress_frame = ttk.LabelFrame(main_frame, text="进度信息", padding="5")
        progress_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        progress_frame.columnconfigure(0, weight=1)
        
        self.progress_var = tk.StringVar(value="准备就绪")
        self.progress_label = ttk.Label(progress_frame, textvariable=self.progress_var)
        self.progress_label.grid(row=0, column=0, sticky=tk.W)
        
        self.progress_bar = ttk.Progressbar(progress_frame, mode='indeterminate')
        self.progress_bar.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(5, 0))
        
        # 日志输出区域
        log_frame = ttk.LabelFrame(main_frame, text="操作日志", padding="5")
        log_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=15, wrap=tk.WORD)
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.log_text.config(state='disabled')
    
    def browse_file(self):
        """浏览并选择Excel文件"""
        filename = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if filename:
            self.file_path_var.set(filename)
    
    def on_log_level_changed(self, event):
        """处理日志级别选择变化"""
        level_text = self.log_level_var.get()
        if level_text == "所有信息":
            self.log_level = LOG_LEVEL['LOG']
        elif level_text == "操作信息":
            self.log_level = LOG_LEVEL['OPER']
        elif level_text == "仅错误":
            self.log_level = LOG_LEVEL['ERR']
    
    def log_message(self, message, level='LOG'):
        """添加日志消息（支持级别过滤）"""
        # 根据当前日志级别决定是否显示
        if level == 'LOG' and self.log_level <= LOG_LEVEL['LOG']:
            self.log_queue.put(message)
        elif level == 'OPER' and self.log_level <= LOG_LEVEL['OPER']:
            self.log_queue.put(message)
        elif level == 'ERR' and self.log_level <= LOG_LEVEL['ERR']:
            self.log_queue.put(message)
    
    def process_log_queue(self):
        """处理日志队列中的消息"""
        try:
            while True:
                message = self.log_queue.get_nowait()
                self.log_text.config(state='normal')
                self.log_text.insert(tk.END, message + '\n')
                self.log_text.see(tk.END)
                self.log_text.config(state='disabled')
        except queue.Empty:
            pass
        
        # 100ms后再次检查队列
        self.root.after(100, self.process_log_queue)
    
    def start_update(self):
        """开始更新基金数据"""
        if not os.path.exists(self.file_path_var.get()):
            messagebox.showerror("错误", "指定的Excel文件不存在！")
            return
        
        self.is_running = True
        self.start_btn.config(state='disabled')
        self.stop_btn.config(state='normal')
        self.browse_btn.config(state='disabled')
        self.progress_bar.start()
        
        # 在新线程中执行更新操作
        self.update_thread = threading.Thread(target=self.update_funds_thread)
        self.update_thread.daemon = True
        self.update_thread.start()
    
    def stop_update(self):
        """停止更新操作"""
        self.is_running = False
        self.start_btn.config(state='normal')
        self.stop_btn.config(state='disabled')
        self.browse_btn.config(state='normal')
        self.progress_bar.stop()
        
        self.log_message("=== 更新操作已停止 ===")
    
    def reset_ui(self):
        """重置界面"""
        self.stop_update()
        
        self.log_text.config(state='normal')
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state='disabled')
        
        self.progress_var.set("准备就绪")
    
    def update_funds_thread(self):
        """在新线程中更新基金数据"""
        try:
            self.log_message("=== 开始更新基金数据 ===")
            
            # 初始化Excel更新器
            self.excel_updater = FundExcelUpdater(self.file_path_var.get())
            
            # 读取Excel文件
            self.log_message("正在读取Excel文件...")
            self.excel_updater.read_excel_fund_values()
            total_funds = len(self.excel_updater.excel_fund_values)
            self.log_message(f"成功读取 {total_funds} 个基金数据")
            
            # 更新基金数据
            successful_updates = 0
            failed_fetches = 0
            
            for i, fund_code in enumerate(sorted(self.excel_updater.excel_fund_values.keys())):
                if not self.is_running:
                    break
                
                # 更新进度
                progress_text = f"处理基金 {fund_code} ({i+1}/{total_funds})"
                self.progress_var.set(progress_text)
                self.log_message(f"\n正在处理基金: {fund_code}")
                
                # 从今天开始逐日往前推最多7天
                found_valid_data = False
                target_date = datetime.today()
                
                for day_offset in range(8):
                    if not self.is_running:
                        break
                    
                    date_str = target_date.strftime("%Y-%m-%d")
                    self.log_message(f"尝试获取 {date_str} 的净值数据 (第{day_offset + 1}次尝试)")
                    
                    success, fund_data = self.fund_data_fetcher.fetch_fund_history_data(fund_code, date_str)
                    
                    if success and fund_data and fund_data.get('net_value') is not None:
                        # 找到有效数据
                        old_net_value, old_percentage = self.excel_updater.excel_fund_values[fund_code]
                        net_value = fund_data['net_value']
                        percentage = fund_data.get('change_percent')
                        
                        # 更新数据
                        self.excel_updater.excel_fund_values[fund_code] = (
                            net_value, 
                            f"{percentage:.2f}%" if percentage is not None else old_percentage
                        )
                        
                        self.log_message(f"✅ 成功更新基金 {fund_code} {date_str} 净值: {old_net_value} -> {net_value}")
                        if percentage is not None:
                            self.log_message(f"   涨跌幅: {old_percentage} -> {percentage:.2f}%")
                        
                        successful_updates += 1
                        found_valid_data = True
                        break
                    else:
                        # 当前日期无数据，往前推一天
                        target_date = target_date - timedelta(days=1)
                
                if not found_valid_data:
                    self.log_message(f"❌ 基金 {fund_code} 最近7天均无有效净值数据")
                    failed_fetches += 1
            
            # 更新Excel文件
            if self.is_running:
                self.log_message("\n正在更新Excel文件...")
                self.excel_updater.update_excel_values()
                self.log_message("✅ Excel文件更新完成")
            
            # 显示结果
            if self.is_running:
                self.log_message(f"\n=== 更新完成 ===")
                self.log_message(f"成功更新: {successful_updates} 个基金")
                self.log_message(f"更新失败: {failed_fetches} 个基金")
                self.progress_var.set(f"更新完成 - 成功: {successful_updates}, 失败: {failed_fetches}")
            
        except Exception as e:
            self.log_message(f"❌ 错误: {str(e)}")
            self.progress_var.set("更新过程中出现错误")
        
        finally:
            # 恢复UI状态
            self.root.after(0, self.stop_update)

def main():
    """主函数"""
    root = tk.Tk()
    app = FundUpdaterGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()